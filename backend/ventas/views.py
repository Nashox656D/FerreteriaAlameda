

from rest_framework import viewsets
from rest_framework.decorators import action
from rest_framework.response import Response
from rest_framework import status
from django.http import HttpResponse
from django.db import transaction
from .models import Cliente, Cotizacion, Pedido, DetallePedido
from .serializers import ClienteSerializer, CotizacionSerializer, PedidoSerializer
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime, date
from io import BytesIO

class ClienteViewSet(viewsets.ModelViewSet):
        queryset = Cliente.objects.all()
        serializer_class = ClienteSerializer
        
        def get_queryset(self):
                if self.request.user.is_superuser:
                        return Cliente.objects.all()
                elif self.request.user.is_authenticated and hasattr(self.request.user, 'empleado'):
                        emp = self.request.user.empleado
                        if emp.cargo in ['Recursos Humanos', 'Admin']:
                                return Cliente.objects.all()
                return Cliente.objects.none()
        
        @action(detail=False, methods=['get'])
        def descargar_excel(self, request):
                from rest_framework.response import Response
                from rest_framework import status
                from rrhh.models import Empleado
                
                if request.user.is_authenticated and hasattr(request.user, 'empleado'):
                        emp = request.user.empleado
                        if emp.cargo == 'Empleado' and not request.user.is_superuser:
                                return Response({'detail': 'No autorizado'}, status=status.HTTP_403_FORBIDDEN)
                
                clientes = Cliente.objects.all()
                
                wb = Workbook()
                ws = wb.active
                ws.title = "Clientes"
                
                header_fill = PatternFill(start_color="4A90E2", end_color="4A90E2", fill_type="solid")
                header_font = Font(bold=True, color="FFFFFF")
                border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
                
                headers = ['ID', 'Nombre', 'RUT', 'Email', 'Teléfono', 'Dirección']
                for col_num, header in enumerate(headers, 1):
                        cell = ws.cell(row=1, column=col_num, value=header)
                        cell.fill = header_fill
                        cell.font = header_font
                        cell.alignment = Alignment(horizontal='center', vertical='center')
                        cell.border = border
                
                for row_num, cliente in enumerate(clientes, 2):
                        row_data = [cliente.id, cliente.nombre, cliente.rut, cliente.email, cliente.telefono, cliente.direccion]
                        for col_num, value in enumerate(row_data, 1):
                                cell = ws.cell(row=row_num, column=col_num, value=value)
                                cell.border = border
                
                column_widths = [8, 30, 15, 30, 15, 40]
                for col_num, width in enumerate(column_widths, 1):
                        ws.column_dimensions[ws.cell(row=1, column=col_num).column_letter].width = width
                
                output = BytesIO()
                wb.save(output)
                output.seek(0)
                
                response = HttpResponse(output.getvalue(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
                response['Content-Disposition'] = f'attachment; filename="Clientes_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx"'
                return response

class CotizacionViewSet(viewsets.ModelViewSet):
        queryset = Cotizacion.objects.all()
        serializer_class = CotizacionSerializer
        
        def get_queryset(self):
                from rrhh.models import Empleado
                
                if self.request.user.is_superuser:
                        return Cotizacion.objects.all()
                elif self.request.user.is_authenticated and hasattr(self.request.user, 'empleado'):
                        emp = self.request.user.empleado
                        if emp.cargo in ['Recursos Humanos', 'Admin']:
                                return Cotizacion.objects.all()
                        else:
                                return Cotizacion.objects.filter(empleado=emp)
                return Cotizacion.objects.none()
        
        def perform_create(self, serializer):
                if not serializer.validated_data.get('empleado') and not serializer.validated_data.get('cliente'):
                        if hasattr(self.request.user, 'empleado'):
                                serializer.save(empleado=self.request.user.empleado)
                        else:
                                serializer.save()
                else:
                        serializer.save()
        
        @action(detail=False, methods=['get'])
        def descargar_excel(self, request):
                from rrhh.models import Empleado
                from rest_framework.response import Response
                from rest_framework import status as http_status
                
                if request.user.is_authenticated and hasattr(request.user, 'empleado'):
                        emp = request.user.empleado
                        if emp.cargo == 'Empleado' and not request.user.is_superuser:
                                return Response({'detail': 'No autorizado'}, status=http_status.HTTP_403_FORBIDDEN)
                
                cotizaciones = self.get_queryset().select_related('cliente', 'empleado')
                
                wb = Workbook()
                ws = wb.active
                ws.title = "Cotizaciones"
                
                headers = ['ID', 'Cliente/Empleado', 'RUT', 'Fecha', 'Fecha Vencimiento', 'Total', 'Estado', 'Pagada', 'Por Pagar']
                ws.append(headers)
                
                header_fill = PatternFill(start_color="4A90E2", end_color="4A90E2", fill_type="solid")
                header_font = Font(bold=True, color="FFFFFF")
                thin_border = Border(
                        left=Side(style='thin'),
                        right=Side(style='thin'),
                        top=Side(style='thin'),
                        bottom=Side(style='thin')
                )
                
                for cell in ws[1]:
                        cell.fill = header_fill
                        cell.font = header_font
                        cell.alignment = Alignment(horizontal='center', vertical='center')
                        cell.border = thin_border
                
                hoy = date.today()
                
                for cotizacion in cotizaciones:
                        por_pagar = "Sí" if not cotizacion.pagada else "No"
                        vencida = ""
                        if cotizacion.fecha_vencimiento and not cotizacion.pagada:
                                if cotizacion.fecha_vencimiento < hoy:
                                        vencida = " (VENCIDA)"
                        
                        nombre = cotizacion.cliente.nombre if cotizacion.cliente else (cotizacion.empleado.nombre if cotizacion.empleado else 'N/A')
                        rut = cotizacion.cliente.rut if cotizacion.cliente else (cotizacion.empleado.rut if cotizacion.empleado else 'N/A')
                        
                        ws.append([
                                cotizacion.id,
                                nombre,
                                rut,
                                cotizacion.fecha.strftime('%Y-%m-%d'),
                                cotizacion.fecha_vencimiento.strftime('%Y-%m-%d') if cotizacion.fecha_vencimiento else 'N/A',
                                float(cotizacion.total),
                                cotizacion.estado,
                                'Sí' if cotizacion.pagada else 'No',
                                por_pagar + vencida
                        ])
                
                for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=9):
                        for cell in row:
                                cell.border = thin_border
                                cell.alignment = Alignment(horizontal='center', vertical='center')
                        
                        pagada_cell = row[7]
                        por_pagar_cell = row[8]
                        
                        if pagada_cell.value == 'Sí':
                                pagada_cell.fill = PatternFill(start_color="C8E6C9", end_color="C8E6C9", fill_type="solid")
                                pagada_cell.font = Font(color="2E7D32", bold=True)
                        else:
                                pagada_cell.fill = PatternFill(start_color="FFCDD2", end_color="FFCDD2", fill_type="solid")
                                pagada_cell.font = Font(color="C62828", bold=True)
                        
                        if "VENCIDA" in str(por_pagar_cell.value):
                                por_pagar_cell.fill = PatternFill(start_color="FFCDD2", end_color="FFCDD2", fill_type="solid")
                                por_pagar_cell.font = Font(color="C62828", bold=True)
                        elif por_pagar_cell.value == "Sí":
                                por_pagar_cell.fill = PatternFill(start_color="FFE082", end_color="FFE082", fill_type="solid")
                                por_pagar_cell.font = Font(color="F57C00", bold=True)
                
                column_widths = [8, 25, 15, 15, 18, 12, 15, 10, 20]
                for col_num, width in enumerate(column_widths, 1):
                        ws.column_dimensions[ws.cell(row=1, column=col_num).column_letter].width = width
                
                output = BytesIO()
                wb.save(output)
                output.seek(0)
                
                response = HttpResponse(
                        output.getvalue(),
                        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
                )
                filename = f"Cotizaciones_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
                response['Content-Disposition'] = f'attachment; filename="{filename}"'
                
                return response

class PedidoViewSet(viewsets.ModelViewSet):
        queryset = Pedido.objects.all()
        serializer_class = PedidoSerializer
        
        def get_queryset(self):
                if self.request.user.is_superuser:
                        return Pedido.objects.all()
                elif self.request.user.is_authenticated and hasattr(self.request.user, 'cliente'):
                        return Pedido.objects.filter(cliente=self.request.user.cliente)
                elif self.request.user.is_authenticated and hasattr(self.request.user, 'empleado'):
                        emp = self.request.user.empleado
                        if emp.cargo in ['Recursos Humanos', 'Admin']:
                                return Pedido.objects.all()
                return Pedido.objects.none()
        
        @action(detail=False, methods=['post'])
        def realizar_compra(self, request):
                from inventario.models import Stock, Producto
                from finanzas.models import Factura
                from decimal import Decimal
                
                if not request.user.is_authenticated or not hasattr(request.user, 'cliente'):
                        return Response({'detail': 'Debe ser un cliente para realizar compras'}, status=status.HTTP_403_FORBIDDEN)
                
                cliente = request.user.cliente
                items = request.data.get('items', [])
                observaciones = request.data.get('observaciones', '')
                generar_factura_pagada = request.data.get('factura_pagada', False)
                
                if not items or not isinstance(items, list):
                        return Response({'detail': 'Debe incluir al menos un producto'}, status=status.HTTP_400_BAD_REQUEST)
                
                try:
                        with transaction.atomic():
                                pedido = Pedido.objects.create(
                                        cliente=cliente,
                                        observaciones=observaciones
                                )
                                
                                total_pedido = Decimal('0')
                                
                                for item in items:
                                        producto_id = item.get('producto_id')
                                        cantidad = item.get('cantidad')
                                        
                                        if not producto_id or not cantidad:
                                                raise ValueError('Debe especificar producto_id y cantidad para cada item')
                                        
                                        try:
                                                cantidad = int(cantidad)
                                                if cantidad <= 0:
                                                        raise ValueError('La cantidad debe ser un número entero positivo')
                                        except (ValueError, TypeError):
                                                raise ValueError(f'Cantidad inválida: {cantidad}. Debe ser un número entero positivo')
                                        
                                        try:
                                                producto = Producto.objects.get(id=producto_id)
                                        except Producto.DoesNotExist:
                                                raise ValueError(f'Producto {producto_id} no encontrado')
                                        
                                        try:
                                                stock = Stock.objects.select_for_update().get(producto=producto)
                                        except Stock.DoesNotExist:
                                                raise ValueError(f'Producto {producto.nombre} sin stock disponible')
                                        
                                        if stock.cantidad < cantidad:
                                                raise ValueError(f'Stock insuficiente para {producto.nombre}. Disponible: {stock.cantidad}, solicitado: {cantidad}')
                                        
                                        stock.cantidad -= cantidad
                                        stock.save()
                                        
                                        precio = producto.precio
                                        subtotal = precio * Decimal(str(cantidad))
                                        
                                        DetallePedido.objects.create(
                                                pedido=pedido,
                                                producto=producto,
                                                cantidad=cantidad,
                                                precio_unitario=precio,
                                                subtotal=subtotal
                                        )
                                        
                                        total_pedido += subtotal
                                
                                pedido.total = total_pedido
                                pedido.save()
                                
                                factura = Factura.objects.create(
                                        cliente=cliente,
                                        monto=total_pedido,
                                        pagada=generar_factura_pagada,
                                        observaciones=f'Factura generada por pedido #{pedido.id}'
                                )
                                
                                return Response({
                                        'detail': 'Compra realizada exitosamente',
                                        'pedido_id': pedido.id,
                                        'factura_id': factura.id,
                                        'total': str(total_pedido)
                                }, status=status.HTTP_201_CREATED)
                                
                except ValueError as e:
                        return Response({'detail': str(e)}, status=status.HTTP_400_BAD_REQUEST)
                except Exception as e:
                        return Response({'detail': f'Error al procesar la compra: {str(e)}'}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
