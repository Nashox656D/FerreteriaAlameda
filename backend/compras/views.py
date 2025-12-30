

from rest_framework import viewsets
from rest_framework.decorators import action
from django.http import HttpResponse
from .models import Proveedor, OrdenCompra, DocumentoCompraInternacional
from .serializers import ProveedorSerializer, OrdenCompraSerializer, DocumentoCompraInternacionalSerializer
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime
from io import BytesIO

class ProveedorViewSet(viewsets.ModelViewSet):
        queryset = Proveedor.objects.all()
        serializer_class = ProveedorSerializer

class OrdenCompraViewSet(viewsets.ModelViewSet):
        queryset = OrdenCompra.objects.all()
        serializer_class = OrdenCompraSerializer
        
        @action(detail=False, methods=['get'])
        def descargar_excel(self, request):
                from rest_framework.response import Response
                from rest_framework import status
                from rrhh.models import Empleado
                
                if request.user.is_authenticated and hasattr(request.user, 'empleado'):
                        emp = request.user.empleado
                        if emp.cargo == 'Empleado' and not request.user.is_superuser:
                                return Response({'detail': 'No autorizado'}, status=status.HTTP_403_FORBIDDEN)
                
                ordenes = OrdenCompra.objects.select_related('proveedor').all()
                
                wb = Workbook()
                ws = wb.active
                ws.title = "Órdenes de Compra"
                
                header_fill = PatternFill(start_color="4A90E2", end_color="4A90E2", fill_type="solid")
                header_font = Font(bold=True, color="FFFFFF")
                border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
                
                headers = ['ID', 'Proveedor', 'RUT', 'Fecha', 'Total', 'Estado', 'Observaciones']
                for col_num, header in enumerate(headers, 1):
                        cell = ws.cell(row=1, column=col_num, value=header)
                        cell.fill = header_fill
                        cell.font = header_font
                        cell.alignment = Alignment(horizontal='center', vertical='center')
                        cell.border = border
                
                for row_num, orden in enumerate(ordenes, 2):
                        row_data = [
                                orden.id,
                                orden.proveedor.nombre,
                                orden.proveedor.rut,
                                orden.fecha.strftime('%Y-%m-%d'),
                                float(orden.total),
                                orden.estado,
                                orden.observaciones or ''
                        ]
                        for col_num, value in enumerate(row_data, 1):
                                cell = ws.cell(row=row_num, column=col_num, value=value)
                                cell.border = border
                
                column_widths = [8, 30, 15, 12, 15, 15, 40]
                for col_num, width in enumerate(column_widths, 1):
                        ws.column_dimensions[ws.cell(row=1, column=col_num).column_letter].width = width
                
                output = BytesIO()
                wb.save(output)
                output.seek(0)
                
                response = HttpResponse(output.getvalue(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
                response['Content-Disposition'] = f'attachment; filename="OrdenesCompra_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx"'
                return response

class DocumentoCompraInternacionalViewSet(viewsets.ModelViewSet):
        queryset = DocumentoCompraInternacional.objects.all()
        serializer_class = DocumentoCompraInternacionalSerializer
