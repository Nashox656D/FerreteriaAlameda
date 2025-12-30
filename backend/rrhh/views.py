from rest_framework import viewsets
from rest_framework.decorators import action
from rest_framework.response import Response
from django.http import HttpResponse
from .models import Empleado, Vacaciones
from .serializers import EmpleadoSerializer, VacacionesSerializer
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime
from io import BytesIO

class EmpleadoViewSet(viewsets.ModelViewSet):
        queryset = Empleado.objects.all()
        serializer_class = EmpleadoSerializer
        
        @action(detail=False, methods=['get'])
        def descargar_excel(self, request):
                from rest_framework.response import Response
                from rest_framework import status
                
                if request.user.is_authenticated and hasattr(request.user, 'empleado'):
                        emp = request.user.empleado
                        if emp.cargo == 'Empleado' and not request.user.is_superuser:
                                return Response({'detail': 'No autorizado'}, status=status.HTTP_403_FORBIDDEN)
                
                empleados = Empleado.objects.all()
                
                wb = Workbook()
                ws = wb.active
                ws.title = "Empleados"
                
                header_fill = PatternFill(start_color="4A90E2", end_color="4A90E2", fill_type="solid")
                header_font = Font(bold=True, color="FFFFFF")
                border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
                
                headers = ['ID', 'Nombre', 'RUT', 'Cargo', 'Email', 'Fecha Ingreso', 'Activo']
                for col_num, header in enumerate(headers, 1):
                        cell = ws.cell(row=1, column=col_num, value=header)
                        cell.fill = header_fill
                        cell.font = header_font
                        cell.alignment = Alignment(horizontal='center', vertical='center')
                        cell.border = border
                
                for row_num, emp in enumerate(empleados, 2):
                        row_data = [
                                emp.id,
                                emp.nombre,
                                emp.rut,
                                emp.cargo,
                                emp.email,
                                emp.fecha_ingreso.strftime('%Y-%m-%d'),
                                'Sí' if emp.activo else 'No'
                        ]
                        for col_num, value in enumerate(row_data, 1):
                                cell = ws.cell(row=row_num, column=col_num, value=value)
                                cell.border = border
                                if col_num == 7:
                                        if emp.activo:
                                                cell.fill = PatternFill(start_color="C8E6C9", end_color="C8E6C9", fill_type="solid")
                                                cell.font = Font(color="2E7D32", bold=True)
                                        else:
                                                cell.fill = PatternFill(start_color="FFCDD2", end_color="FFCDD2", fill_type="solid")
                                                cell.font = Font(color="C62828", bold=True)
                
                column_widths = [8, 30, 15, 20, 30, 15, 10]
                for col_num, width in enumerate(column_widths, 1):
                        ws.column_dimensions[ws.cell(row=1, column=col_num).column_letter].width = width
                
                output = BytesIO()
                wb.save(output)
                output.seek(0)
                
                response = HttpResponse(output.getvalue(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
                response['Content-Disposition'] = f'attachment; filename="Empleados_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx"'
                return response

class VacacionesViewSet(viewsets.ModelViewSet):
        queryset = Vacaciones.objects.all()
        serializer_class = VacacionesSerializer
        
        def get_queryset(self):
                if self.request.user.is_superuser:
                        return Vacaciones.objects.all()
                elif self.request.user.is_authenticated and hasattr(self.request.user, 'empleado'):
                        emp = self.request.user.empleado
                        if emp.cargo in ['Recursos Humanos', 'Admin']:
                                return Vacaciones.objects.all()
                        else:
                                return Vacaciones.objects.filter(empleado=emp)
                return Vacaciones.objects.none()
        
        def perform_create(self, serializer):
                from rest_framework.exceptions import PermissionDenied, ValidationError
                
                if self.request.user.is_authenticated and hasattr(self.request.user, 'empleado'):
                        emp = self.request.user.empleado
                        if emp.cargo in ['Recursos Humanos', 'Admin']:
                                raise PermissionDenied('Los administradores no pueden crear solicitudes de vacaciones. Solo pueden aprobar o rechazar.')
                        
                        serializer.save(empleado=emp, estado='PENDIENTE', aprobada=False)
                else:
                        raise ValidationError({'empleado': 'Debe ser un empleado para solicitar vacaciones.'})
        
        @action(detail=True, methods=['post'])
        def aprobar(self, request, pk=None):
                from rest_framework import status
                from django.utils import timezone
                
                if not request.user.is_superuser and not (hasattr(request.user, 'empleado') and request.user.empleado.cargo in ['Recursos Humanos', 'Admin']):
                        return Response({'detail': 'No tiene permisos para aprobar vacaciones'}, status=status.HTTP_403_FORBIDDEN)
                
                vacacion = self.get_object()
                
                if vacacion.estado != 'PENDIENTE':
                        return Response({'detail': f'No se puede aprobar una vacación en estado {vacacion.estado}'}, status=status.HTTP_400_BAD_REQUEST)
                
                comentario = request.data.get('comentario_admin', '')
                
                vacacion.estado = 'APROBADA'
                vacacion.aprobada = True
                vacacion.aprobada_por = request.user
                vacacion.fecha_respuesta = timezone.now()
                vacacion.comentario_admin = comentario
                vacacion.save()
                
                serializer = self.get_serializer(vacacion)
                return Response(serializer.data)
        
        @action(detail=True, methods=['post'])
        def rechazar(self, request, pk=None):
                from rest_framework import status
                from django.utils import timezone
                
                if not request.user.is_superuser and not (hasattr(request.user, 'empleado') and request.user.empleado.cargo in ['Recursos Humanos', 'Admin']):
                        return Response({'detail': 'No tiene permisos para rechazar vacaciones'}, status=status.HTTP_403_FORBIDDEN)
                
                vacacion = self.get_object()
                
                if vacacion.estado != 'PENDIENTE':
                        return Response({'detail': f'No se puede rechazar una vacación en estado {vacacion.estado}'}, status=status.HTTP_400_BAD_REQUEST)
                
                comentario = request.data.get('comentario_admin', '')
                
                vacacion.estado = 'RECHAZADA'
                vacacion.aprobada = False
                vacacion.aprobada_por = request.user
                vacacion.fecha_respuesta = timezone.now()
                vacacion.comentario_admin = comentario
                vacacion.save()
                
                serializer = self.get_serializer(vacacion)
                return Response(serializer.data)
        
        @action(detail=False, methods=['get'])
        def descargar_excel(self, request):
                from rest_framework.response import Response
                from rest_framework import status as http_status
                
                vacaciones = self.get_queryset().select_related('empleado')
                
                wb = Workbook()
                ws = wb.active
                ws.title = "Vacaciones"
                
                header_fill = PatternFill(start_color="1E88E5", end_color="1E88E5", fill_type="solid")
                header_font = Font(color="FFFFFF", bold=True, size=12)
                border = Border(
                        left=Side(style='thin'),
                        right=Side(style='thin'),
                        top=Side(style='thin'),
                        bottom=Side(style='thin')
                )
                
                headers = ['ID', 'Empleado', 'RUT', 'Cargo', 'Fecha Inicio', 'Fecha Fin', 'Días', 'Estado', 'Aprobada']
                for col_num, header in enumerate(headers, 1):
                        cell = ws.cell(row=1, column=col_num, value=header)
                        cell.fill = header_fill
                        cell.font = header_font
                        cell.alignment = Alignment(horizontal='center', vertical='center')
                        cell.border = border
                
                for row_num, vac in enumerate(vacaciones, 2):
                        dias = (vac.fecha_fin - vac.fecha_inicio).days + 1
                        estado = vac.get_estado_display()
                        
                        row_data = [
                                vac.id,
                                vac.empleado.nombre,
                                vac.empleado.rut,
                                vac.empleado.cargo,
                                vac.fecha_inicio.strftime('%d/%m/%Y'),
                                vac.fecha_fin.strftime('%d/%m/%Y'),
                                dias,
                                estado,
                                "Sí" if vac.aprobada else "No"
                        ]
                        
                        for col_num, value in enumerate(row_data, 1):
                                cell = ws.cell(row=row_num, column=col_num, value=value)
                                cell.border = border
                                cell.alignment = Alignment(horizontal='center' if col_num in [1, 7, 8, 9] else 'left', vertical='center')
                                
                                if col_num == 8:
                                        if vac.aprobada:
                                                cell.fill = PatternFill(start_color="C8E6C9", end_color="C8E6C9", fill_type="solid")
                                                cell.font = Font(color="2E7D32", bold=True)
                                        else:
                                                cell.fill = PatternFill(start_color="FFE082", end_color="FFE082", fill_type="solid")
                                                cell.font = Font(color="F57C00", bold=True)
                
                column_widths = [8, 25, 15, 20, 15, 15, 10, 15, 12]
                for col_num, width in enumerate(column_widths, 1):
                        ws.column_dimensions[ws.cell(row=1, column=col_num).column_letter].width = width
                
                output = BytesIO()
                wb.save(output)
                output.seek(0)
                
                response = HttpResponse(
                        output.getvalue(),
                        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
                )
                filename = f"Vacaciones_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
                response['Content-Disposition'] = f'attachment; filename="{filename}"'
                
                return response
