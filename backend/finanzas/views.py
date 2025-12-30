

from rest_framework import viewsets
from rest_framework.decorators import action
from django.http import HttpResponse
from .models import Factura, Comision
from .serializers import FacturaSerializer, ComisionSerializer
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime, date
from io import BytesIO

class FacturaViewSet(viewsets.ModelViewSet):
        queryset = Factura.objects.all()
        serializer_class = FacturaSerializer
        
        @action(detail=False, methods=['get'])
        def descargar_excel(self, request):
                facturas = Factura.objects.all().prefetch_related('lineas', 'cliente')
                
                wb = Workbook()
                ws = wb.active
                ws.title = "Facturas"
                
                headers = ['ID', 'Número', 'Cliente', 'Fecha', 'Fecha Vencimiento', 'Monto', 'Pagada', 'Por Pagar']
                ws.append(headers)
                
                header_fill = PatternFill(start_color="4A90E2", end_color="4A90E2", fill_type="solid")
                header_font = Font(bold=True, color="FFFFFF")
                thin_border = Border(
                        left=Side(style='thin'),
                        right=Side(style='thin'),
                        top=Side(style='thin'),
                        bottom=Side(style='thin')
                )
                
                for cell in ws[1]:
                        cell.fill = header_fill
                        cell.font = header_font
                        cell.alignment = Alignment(horizontal='center', vertical='center')
                        cell.border = thin_border
                
                hoy = date.today()
                
                for factura in facturas:
                        por_pagar = "Sí" if not factura.pagada else "No"
                        vencida = ""
                        if factura.fecha_vencimiento and not factura.pagada:
                                if factura.fecha_vencimiento < hoy:
                                        vencida = " (VENCIDA)"
                        
                        ws.append([
                                factura.id,
                                factura.numero,
                                factura.cliente.nombre if factura.cliente else 'Consumidor',
                                factura.fecha.strftime('%Y-%m-%d'),
                                factura.fecha_vencimiento.strftime('%Y-%m-%d') if factura.fecha_vencimiento else 'N/A',
                                float(factura.monto),
                                'Sí' if factura.pagada else 'No',
                                por_pagar + vencida
                        ])
                
                for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=8):
                        for cell in row:
                                cell.border = thin_border
                                cell.alignment = Alignment(horizontal='center', vertical='center')
                        
                        pagada_cell = row[6]
                        por_pagar_cell = row[7]
                        
                        if pagada_cell.value == 'Sí':
                                pagada_cell.fill = PatternFill(start_color="C8E6C9", end_color="C8E6C9", fill_type="solid")
                                pagada_cell.font = Font(color="2E7D32", bold=True)
                        else:
                                pagada_cell.fill = PatternFill(start_color="FFCDD2", end_color="FFCDD2", fill_type="solid")
                                pagada_cell.font = Font(color="C62828", bold=True)
                        
                        if "VENCIDA" in str(por_pagar_cell.value):
                                por_pagar_cell.fill = PatternFill(start_color="FFCDD2", end_color="FFCDD2", fill_type="solid")
                                por_pagar_cell.font = Font(color="C62828", bold=True)
                        elif por_pagar_cell.value == "Sí":
                                por_pagar_cell.fill = PatternFill(start_color="FFE082", end_color="FFE082", fill_type="solid")
                                por_pagar_cell.font = Font(color="F57C00", bold=True)
                
                column_widths = [8, 15, 25, 15, 18, 12, 10, 20]
                for col_num, width in enumerate(column_widths, 1):
                        ws.column_dimensions[ws.cell(row=1, column=col_num).column_letter].width = width
                
                output = BytesIO()
                wb.save(output)
                output.seek(0)
                
                response = HttpResponse(
                        output.getvalue(),
                        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
                )
                filename = f"Facturas_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
                response['Content-Disposition'] = f'attachment; filename="{filename}"'
                
                return response

        @action(detail=False, methods=['get'])
        def descargar_excel_detalle(self, request):
                """Exporta un Excel con dos hojas: resumen de facturas y detalle por línea."""
                facturas = Factura.objects.all().prefetch_related('lineas', 'cliente')

                wb = Workbook()

                # Hoja 1: Resumen
                ws1 = wb.active
                ws1.title = "Facturas_Resumen"
                headers = ['ID', 'Número', 'Cliente', 'Fecha', 'Fecha Vencimiento', 'Monto', 'Pagada']
                ws1.append(headers)

                header_fill = PatternFill(start_color="4A90E2", end_color="4A90E2", fill_type="solid")
                header_font = Font(bold=True, color="FFFFFF")
                thin_border = Border(
                        left=Side(style='thin'),
                        right=Side(style='thin'),
                        top=Side(style='thin'),
                        bottom=Side(style='thin')
                )

                for cell in ws1[1]:
                        cell.fill = header_fill
                        cell.font = header_font
                        cell.alignment = Alignment(horizontal='center', vertical='center')
                        cell.border = thin_border

                for factura in facturas:
                        cliente_nombre = factura.cliente.nombre if factura.cliente else 'Consumidor'
                        ws1.append([
                                factura.id,
                                factura.numero,
                                cliente_nombre,
                                factura.fecha.strftime('%Y-%m-%d'),
                                factura.fecha_vencimiento.strftime('%Y-%m-%d') if factura.fecha_vencimiento else 'N/A',
                                float(factura.monto),
                                'Sí' if factura.pagada else 'No'
                        ])

                # Hoja 2: Detalle de líneas
                ws2 = wb.create_sheet(title='Facturas_Detalle')
                headers2 = ['Factura ID', 'Número', 'Producto SKU', 'Producto', 'Descripción', 'Cantidad', 'Precio Unitario', 'Subtotal']
                ws2.append(headers2)
                for cell in ws2[1]:
                        cell.fill = header_fill
                        cell.font = header_font
                        cell.alignment = Alignment(horizontal='center', vertical='center')
                        cell.border = thin_border

                for factura in facturas:
                        for linea in factura.lineas.all():
                                sku = linea.producto.sku if linea.producto else ''
                                nombre = linea.producto.nombre if linea.producto else (linea.descripcion or '')
                                ws2.append([
                                        factura.id,
                                        factura.numero,
                                        sku,
                                        nombre,
                                        linea.descripcion,
                                        linea.cantidad,
                                        float(linea.precio_unitario),
                                        float(linea.subtotal)
                                ])

                # Ajustar anchos de columnas (ambas hojas)
                for ws in (ws1, ws2):
                        for col in ws.columns:
                                max_length = 0
                                column = col[0].column_letter
                                for cell in col:
                                        try:
                                                val = str(cell.value)
                                                if len(val) > max_length:
                                                        max_length = len(val)
                                        except Exception:
                                                pass
                                adjusted_width = (max_length + 2) if max_length > 0 else 8
                                ws.column_dimensions[column].width = adjusted_width

                output = BytesIO()
                wb.save(output)
                output.seek(0)

                response = HttpResponse(
                        output.getvalue(),
                        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
                )
                filename = f"Facturas_Detalle_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
                response['Content-Disposition'] = f'attachment; filename="{filename}"'

                return response

class ComisionViewSet(viewsets.ModelViewSet):
        queryset = Comision.objects.all()
        serializer_class = ComisionSerializer
