

from rest_framework import viewsets
from rest_framework.decorators import action
from django.http import HttpResponse
from .models import Producto, Stock, MovimientoStock, ImagenProducto
from .serializers import ProductoSerializer, StockSerializer, MovimientoStockSerializer, ImagenProductoSerializer
from rest_framework.response import Response
from rest_framework import status
from rest_framework.permissions import IsAuthenticated
from core.permissions import IsEmpleadoReadOnly
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime
from io import BytesIO

def _user_is_empleado(user):
        # kept as compatibility helper (used in some custom flows)
        try:
                from rrhh.models import Empleado
                emp = Empleado.objects.filter(user=user).first()
                return emp is not None and emp.cargo == 'Empleado'
        except Exception:
                return False

class ProductoViewSet(viewsets.ModelViewSet):
        queryset = Producto.objects.all()
        serializer_class = ProductoSerializer
        permission_classes = [IsAuthenticated, IsEmpleadoReadOnly]

        # The permission class centralizes the "Empleado" read-only rule.
        # Keep method overrides to return friendly 403 when needed (defensive), but
        # the IsEmpleadoReadOnly will already block unsafe methods for employees.

        def create(self, request, *args, **kwargs):
                if _user_is_empleado(request.user) and not request.user.is_superuser:
                        return Response({'detail': 'No autorizado'}, status=status.HTTP_403_FORBIDDEN)
                serializer = self.get_serializer(data=request.data)
                serializer.is_valid(raise_exception=True)
                producto = serializer.save()
                # Log movimiento: producto creado
                try:
                        MovimientoStock.objects.create(
                                producto=producto,
                                stock=None,
                                tipo='producto_creado',
                                cantidad=0,
                                usuario=request.user if request.user.is_authenticated else None,
                                nota=f'Producto creado: {producto.sku} - {producto.nombre}'
                        )
                except Exception:
                        pass
                headers = self.get_success_headers(serializer.data)
                return Response(serializer.data, status=status.HTTP_201_CREATED, headers=headers)

        def update(self, request, *args, **kwargs):
                if _user_is_empleado(request.user) and not request.user.is_superuser:
                        return Response({'detail': 'No autorizado'}, status=status.HTTP_403_FORBIDDEN)
                return super().update(request, *args, **kwargs)

        def destroy(self, request, *args, **kwargs):
                if _user_is_empleado(request.user) and not request.user.is_superuser:
                        return Response({'detail': 'No autorizado'}, status=status.HTTP_403_FORBIDDEN)
                instance = self.get_object()
                sku = instance.sku
                nombre = instance.nombre
                # Log movimiento: producto eliminado
                try:
                        MovimientoStock.objects.create(
                                producto=instance,
                                stock=None,
                                tipo='producto_eliminado',
                                cantidad=0,
                                usuario=request.user if request.user.is_authenticated else None,
                                nota=f'Producto eliminado: {sku} - {nombre}'
                        )
                except Exception:
                        pass
                return super().destroy(request, *args, **kwargs)
        
        @action(detail=False, methods=['get'])
        def descargar_excel(self, request):
                # keep export restricted to non-empleado roles
                if _user_is_empleado(request.user) and not request.user.is_superuser:
                        return Response({'detail': 'No autorizado'}, status=status.HTTP_403_FORBIDDEN)
                productos = Producto.objects.all()
                
                wb = Workbook()
                ws = wb.active
                ws.title = "Productos"
                
                header_fill = PatternFill(start_color="4A90E2", end_color="4A90E2", fill_type="solid")
                header_font = Font(bold=True, color="FFFFFF")
                border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
                
                headers = ['ID', 'SKU', 'Nombre', 'Marca', 'Descripción', 'Precio']
                for col_num, header in enumerate(headers, 1):
                        cell = ws.cell(row=1, column=col_num, value=header)
                        cell.fill = header_fill
                        cell.font = header_font
                        cell.alignment = Alignment(horizontal='center', vertical='center')
                        cell.border = border
                
                for row_num, prod in enumerate(productos, 2):
                        row_data = [prod.id, prod.sku, prod.nombre, prod.marca or '', prod.descripcion, float(prod.precio)]
                        for col_num, value in enumerate(row_data, 1):
                                cell = ws.cell(row=row_num, column=col_num, value=value)
                                cell.border = border
                
                column_widths = [8, 15, 30, 20, 40, 12]
                for col_num, width in enumerate(column_widths, 1):
                        ws.column_dimensions[ws.cell(row=1, column=col_num).column_letter].width = width
                
                output = BytesIO()
                wb.save(output)
                output.seek(0)
                
                response = HttpResponse(output.getvalue(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
                filename = f'Productos_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
                response['Content-Disposition'] = f'attachment; filename="{filename}"'
                try:
                        from urllib.parse import quote
                        response['Content-Disposition'] += f"; filename*=UTF-8''{quote(filename)}"
                except Exception:
                        pass
                return response

class StockViewSet(viewsets.ModelViewSet):
        queryset = Stock.objects.all()
        serializer_class = StockSerializer
        permission_classes = [IsAuthenticated, IsEmpleadoReadOnly]

        def create(self, request, *args, **kwargs):
                # block creation for plain 'Empleado' cargo (defensive; permission also enforces)
                if _user_is_empleado(request.user) and not request.user.is_superuser:
                        return Response({'detail': 'No autorizado'}, status=status.HTTP_403_FORBIDDEN)
                
                # Accept producto as either ID (int) or SKU (string)
                producto_input = request.data.get('producto')
                cantidad = int(request.data.get('cantidad', 0))
                
                if not producto_input or cantidad <= 0:
                        return Response({'error': 'Datos inválidos'}, status=status.HTTP_400_BAD_REQUEST)
                
                try:
                        # Try to get producto by ID first (if numeric), then by SKU
                        if isinstance(producto_input, int) or (isinstance(producto_input, str) and producto_input.isdigit()):
                                producto = Producto.objects.get(id=int(producto_input))
                        else:
                                producto = Producto.objects.get(sku=producto_input)
                except Producto.DoesNotExist:
                        return Response({'error': f'No existe un producto con ID/SKU: {producto_input}'}, status=status.HTTP_404_NOT_FOUND)
                except (ValueError, TypeError):
                        return Response({'error': 'Formato inválido para producto'}, status=status.HTTP_400_BAD_REQUEST)

                stock_qs = Stock.objects.filter(producto=producto)
                if stock_qs.exists():
                        stock = stock_qs.first()
                        old_cantidad = stock.cantidad
                        stock.cantidad += cantidad
                        stock.save()
                        # Log movimiento: agregar
                        try:
                                MovimientoStock.objects.create(
                                        producto=producto,
                                        stock=stock,
                                        tipo='agregar',
                                        cantidad=cantidad,
                                        usuario=request.user if request.user.is_authenticated else None,
                                        nota=f'Carga por SKU {producto.sku} (antes {old_cantidad} -> ahora {stock.cantidad})'
                                )
                        except Exception as e:
                                import traceback
                                print(f"Error al crear MovimientoStock: {e}")
                                traceback.print_exc()
                        serializer = self.get_serializer(stock)
                        return Response(serializer.data)
                else:
                        # create new stock record and log movimiento
                        request.data['producto'] = producto.id
                        serializer = self.get_serializer(data=request.data)
                        serializer.is_valid(raise_exception=True)
                        stock = serializer.save()
                        try:
                                MovimientoStock.objects.create(
                                        producto=producto,
                                        stock=stock,
                                        tipo='agregar',
                                        cantidad=cantidad,
                                        usuario=request.user if request.user.is_authenticated else None,
                                        nota=f'Nuevo stock creado para SKU {producto.sku} con cantidad {cantidad}'
                                )
                        except Exception as e:
                                import traceback
                                print(f"Error al crear MovimientoStock en nuevo stock: {e}")
                                traceback.print_exc()
                        return Response(self.get_serializer(stock).data, status=status.HTTP_201_CREATED)

        def update(self, request, *args, **kwargs):
                # handle delta logging on update
                if _user_is_empleado(request.user) and not request.user.is_superuser:
                        return Response({'detail': 'No autorizado'}, status=status.HTTP_403_FORBIDDEN)
                partial = kwargs.pop('partial', False)
                instance = self.get_object()
                old_cantidad = instance.cantidad
                serializer = self.get_serializer(instance, data=request.data, partial=partial)
                serializer.is_valid(raise_exception=True)
                stock = serializer.save()
                new_cantidad = stock.cantidad
                delta = new_cantidad - old_cantidad
                if delta != 0:
                        tipo = 'agregar' if delta > 0 else 'restar'
                        try:
                                MovimientoStock.objects.create(
                                        producto=stock.producto,
                                        stock=stock,
                                        tipo=tipo,
                                        cantidad=abs(delta),
                                        usuario=request.user if request.user.is_authenticated else None,
                                        nota=f'Actualización de stock (antes {old_cantidad} -> ahora {new_cantidad})'
                                )
                        except Exception:
                                pass
                return Response(self.get_serializer(stock).data)
        
        @action(detail=True, methods=['post'])
        def cambiar_estado(self, request, pk=None):
                """Cambiar estado del stock y registrar auditoría."""
                if _user_is_empleado(request.user) and not request.user.is_superuser:
                        return Response({'detail': 'No autorizado'}, status=status.HTTP_403_FORBIDDEN)
                
                stock = self.get_object()
                nuevo_estado = request.data.get('estado')
                cantidad = int(request.data.get('cantidad', 0))  # Cantidad a cambiar
                motivo = request.data.get('motivo', '')
                
                if not nuevo_estado:
                        return Response({'error': 'Estado requerido'}, status=status.HTTP_400_BAD_REQUEST)
                
                # Validar que el nuevo estado es válido
                estados_validos = [choice[0] for choice in Stock.ESTADO_CHOICES]
                if nuevo_estado not in estados_validos:
                        return Response({'error': f'Estado inválido: {nuevo_estado}'}, status=status.HTTP_400_BAD_REQUEST)
                
                # Si no especifica cantidad, usar toda la cantidad del stock actual
                if cantidad <= 0:
                        cantidad = stock.cantidad
                
                # Validar que no intente cambiar más cantidad de la que tiene
                if cantidad > stock.cantidad:
                        return Response({'error': f'No puedes cambiar {cantidad} unidades cuando solo hay {stock.cantidad}'}, status=status.HTTP_400_BAD_REQUEST)
                
                estado_anterior = stock.estado
                
                # Reducir cantidad del stock original
                stock.cantidad -= cantidad
                
                # Si la cantidad llega a 0, eliminar el registro; si no, guardar
                if stock.cantidad == 0:
                        stock.delete()
                else:
                        stock.save()
                
                # Crear o actualizar stock con el nuevo estado
                stock_nuevo, created = Stock.objects.get_or_create(
                        producto=stock.producto,
                        estado=nuevo_estado,
                        defaults={'cantidad': 0, 'ubicacion': stock.ubicacion}
                )
                stock_nuevo.cantidad += cantidad
                stock_nuevo.motivo_defecto = motivo
                from django.utils import timezone
                stock_nuevo.fecha_defecto = timezone.now()
                stock_nuevo.usuario_reporte = request.user if request.user.is_authenticated else None
                stock_nuevo.save()
                
                # Log movimiento de cambio de estado
                try:
                        MovimientoStock.objects.create(
                                producto=stock_nuevo.producto,
                                stock=stock_nuevo,
                                tipo='cambio_estado',
                                cantidad=cantidad,
                                usuario=request.user if request.user.is_authenticated else None,
                                nota=f'Cambio de estado: {cantidad} unid. {estado_anterior} → {nuevo_estado}. Motivo: {motivo}'
                        )
                except Exception:
                        pass
                
                return Response(self.get_serializer(stock_nuevo).data)
        
        @action(detail=False, methods=['post'])
        def recibir_devolucion(self, request):
                """Registrar devolución de producto y aumentar stock."""
                if _user_is_empleado(request.user) and not request.user.is_superuser:
                        return Response({'detail': 'No autorizado'}, status=status.HTTP_403_FORBIDDEN)
                
                producto_input = request.data.get('producto')
                cantidad = int(request.data.get('cantidad', 0))
                estado_devolucion = request.data.get('estado', 'apto')  # apto o defectuoso
                motivo = request.data.get('motivo', '')
                
                if not producto_input or cantidad <= 0:
                        return Response({'error': 'Datos inválidos'}, status=status.HTTP_400_BAD_REQUEST)
                
                # Obtener producto
                try:
                        if isinstance(producto_input, int) or (isinstance(producto_input, str) and producto_input.isdigit()):
                                producto = Producto.objects.get(id=int(producto_input))
                        else:
                                producto = Producto.objects.get(sku=producto_input)
                except Producto.DoesNotExist:
                        return Response({'error': f'No existe un producto con ID/SKU: {producto_input}'}, status=status.HTTP_404_NOT_FOUND)
                
                # Validar estado de devolución
                estados_validos = [choice[0] for choice in Stock.ESTADO_CHOICES]
                if estado_devolucion not in estados_validos:
                        estado_devolucion = 'apto'
                
                # Crear o actualizar stock con el estado de devolución
                stock, created = Stock.objects.get_or_create(
                        producto=producto,
                        estado=estado_devolucion,
                        defaults={'cantidad': 0, 'ubicacion': ''}
                )
                stock.cantidad += cantidad
                stock.save()
                
                # Log movimiento de devolución
                try:
                        MovimientoStock.objects.create(
                                producto=producto,
                                stock=stock,
                                tipo='devolucion',
                                cantidad=cantidad,
                                usuario=request.user if request.user.is_authenticated else None,
                                nota=f'Devolución: {cantidad} unid. en estado {estado_devolucion}. Motivo: {motivo}'
                        )
                except Exception:
                        pass
                
                return Response(self.get_serializer(stock).data, status=status.HTTP_201_CREATED)
        def descargar_excel(self, request):
                if _user_is_empleado(request.user) and not request.user.is_superuser:
                        return Response({'detail': 'No autorizado'}, status=status.HTTP_403_FORBIDDEN)
                stock = Stock.objects.select_related('producto').all()
                
                wb = Workbook()
                ws = wb.active
                ws.title = "Stock"
                
                header_fill = PatternFill(start_color="4A90E2", end_color="4A90E2", fill_type="solid")
                header_font = Font(bold=True, color="FFFFFF")
                border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
                
                headers = ['ID', 'SKU Producto', 'Nombre Producto', 'Cantidad', 'Ubicación']
                for col_num, header in enumerate(headers, 1):
                        cell = ws.cell(row=1, column=col_num, value=header)
                        cell.fill = header_fill
                        cell.font = header_font
                        cell.alignment = Alignment(horizontal='center', vertical='center')
                        cell.border = border
                
                for row_num, s in enumerate(stock, 2):
                        row_data = [s.id, s.producto.sku, s.producto.nombre, s.cantidad, s.ubicacion or '']
                        for col_num, value in enumerate(row_data, 1):
                                cell = ws.cell(row=row_num, column=col_num, value=value)
                                cell.border = border
                                if col_num == 4:
                                        cell.alignment = Alignment(horizontal='center')
                
                column_widths = [8, 15, 30, 12, 20]
                for col_num, width in enumerate(column_widths, 1):
                        ws.column_dimensions[ws.cell(row=1, column=col_num).column_letter].width = width
                
                output = BytesIO()
                wb.save(output)
                output.seek(0)
                
                response = HttpResponse(output.getvalue(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
                filename = f'Stock_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
                response['Content-Disposition'] = f'attachment; filename="{filename}"'
                try:
                        from urllib.parse import quote
                        response['Content-Disposition'] += f"; filename*=UTF-8''{quote(filename)}"
                except Exception:
                        pass
                return response

        @action(detail=False, methods=['get'])
        def descargar_movimientos(self, request):
                # export movimientos to excel; optional filters: producto, desde, hasta
                if _user_is_empleado(request.user) and not request.user.is_superuser:
                        return Response({'detail': 'No autorizado'}, status=status.HTTP_403_FORBIDDEN)
                qs = MovimientoStock.objects.select_related('producto', 'stock', 'usuario').all().order_by('-creado')
                producto_sku = request.query_params.get('producto')
                desde = request.query_params.get('desde')
                hasta = request.query_params.get('hasta')
                
                filtros_aplicados = []
                if producto_sku:
                        qs = qs.filter(producto__sku=producto_sku)
                        filtros_aplicados.append(f'SKU: {producto_sku}')
                try:
                        if desde:
                                desde_dt = datetime.fromisoformat(desde)
                                qs = qs.filter(creado__gte=desde_dt)
                                filtros_aplicados.append(f'Desde: {desde_dt.strftime("%Y-%m-%d")}')
                        if hasta:
                                hasta_dt = datetime.fromisoformat(hasta)
                                qs = qs.filter(creado__lte=hasta_dt)
                                filtros_aplicados.append(f'Hasta: {hasta_dt.strftime("%Y-%m-%d")}')
                except Exception:
                        pass

                wb = Workbook()
                ws = wb.active
                ws.title = 'Movimientos'
                
                # Estilos
                header_fill = PatternFill(start_color='4A90E2', end_color='4A90E2', fill_type='solid')
                header_font = Font(bold=True, color='FFFFFF', size=12)
                border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
                border_thick = Border(left=Side(style='medium'), right=Side(style='medium'), top=Side(style='medium'), bottom=Side(style='medium'))
                
                # Colores por tipo de movimiento
                tipo_colors = {
                        'agregar': 'E8F5E9',      # Verde claro
                        'restar': 'FFEBEE',       # Rojo claro
                        'venta': 'FFF3E0',        # Naranja claro
                        'producto_creado': 'E3F2FD',  # Azul claro
                        'producto_eliminado': 'F3E5F5', # Púrpura claro
                        'cambio_estado': 'FFF9C4', # Amarillo claro
                        'reclamo_proveedor': 'E0F2F1', # Teal claro
                        'dado_baja': 'ECEFF1',    # Gris azulado
                        'devolucion': 'C8E6C9',   # Verde
                }
                
                # Título y filtros (filas 1-3)
                ws.merge_cells('A1:H1')
                title_cell = ws['A1']
                title_cell.value = '📊 REPORTE DE MOVIMIENTOS DE INVENTARIO'
                title_cell.font = Font(bold=True, size=14, color='FFFFFF')
                title_cell.fill = PatternFill(start_color='1565C0', end_color='1565C0', fill_type='solid')
                title_cell.alignment = Alignment(horizontal='center', vertical='center')
                ws.row_dimensions[1].height = 25
                
                # Información de generación
                ws['A2'] = f'Generado: {datetime.now().strftime("%Y-%m-%d %H:%M:%S")}'
                ws['A2'].font = Font(italic=True, size=10, color='666666')
                
                # Filtros aplicados
                if filtros_aplicados:
                        ws['A3'] = f'Filtros: {" | ".join(filtros_aplicados)}'
                        ws['A3'].font = Font(italic=True, size=10, color='1976D2')
                
                # Encabezados (fila 5)
                headers = ['ID', 'Fecha/Hora', 'SKU', 'Producto', 'Marca', 'Tipo Movimiento', 'Cantidad', 'Usuario', 'Nota']
                for col_num, header in enumerate(headers, 1):
                        cell = ws.cell(row=5, column=col_num, value=header)
                        cell.fill = header_fill
                        cell.font = header_font
                        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                        cell.border = border
                ws.row_dimensions[5].height = 20

                # Datos de movimientos
                row = 6
                totales_por_tipo = {}
                
                for m in qs:
                        tipo_key = m.tipo
                        row_data = [
                                m.id,
                                m.creado.strftime('%Y-%m-%d %H:%M:%S'),
                                m.producto.sku if m.producto else '',
                                m.producto.nombre if m.producto else '',
                                m.producto.marca if m.producto and m.producto.marca else '',
                                m.get_tipo_display(),
                                m.cantidad,
                                m.usuario.username if m.usuario else 'Sistema',
                                m.nota
                        ]
                        
                        for col_num, value in enumerate(row_data, 1):
                                cell = ws.cell(row=row, column=col_num, value=value)
                                cell.border = border
                                
                                # Colorear según tipo de movimiento
                                if tipo_colors.get(tipo_key):
                                        cell.fill = PatternFill(start_color=tipo_colors[tipo_key], end_color=tipo_colors[tipo_key], fill_type='solid')
                                
                                # Alineación
                                if col_num in [1, 7]:  # ID y Cantidad
                                        cell.alignment = Alignment(horizontal='center')
                                elif col_num == 2:  # Fecha
                                        cell.alignment = Alignment(horizontal='center')
                                        cell.number_format = 'YYYY-MM-DD HH:MM:SS'
                                else:
                                        cell.alignment = Alignment(horizontal='left', wrap_text=True)
                        
                        # Contar totales por tipo
                        if tipo_key not in totales_por_tipo:
                                totales_por_tipo[tipo_key] = 0
                        totales_por_tipo[tipo_key] += m.cantidad
                        
                        row += 1

                # Resumen final
                summary_row = row + 2
                ws.merge_cells(f'A{summary_row}:E{summary_row}')
                summary_cell = ws[f'A{summary_row}']
                summary_cell.value = '📈 RESUMEN POR TIPO DE MOVIMIENTO'
                summary_cell.font = Font(bold=True, size=12, color='FFFFFF')
                summary_cell.fill = PatternFill(start_color='1565C0', end_color='1565C0', fill_type='solid')
                summary_cell.alignment = Alignment(horizontal='center', vertical='center')
                
                # Encabezados del resumen
                summary_row += 1
                summary_headers = ['Tipo Movimiento', 'Cantidad Total', 'Color']
                for col_num, header in enumerate(summary_headers, 1):
                        cell = ws.cell(row=summary_row, column=col_num, value=header)
                        cell.fill = header_fill
                        cell.font = header_font
                        cell.border = border
                        cell.alignment = Alignment(horizontal='center')
                
                # Datos del resumen
                summary_row += 1
                for tipo_key, cantidad_total in sorted(totales_por_tipo.items()):
                        ws.cell(row=summary_row, column=1, value=dict(MovimientoStock.TIPO_CHOICES).get(tipo_key, tipo_key))
                        ws.cell(row=summary_row, column=2, value=cantidad_total)
                        ws.cell(row=summary_row, column=2).alignment = Alignment(horizontal='center')
                        
                        # Color indicador
                        for col in range(1, 4):
                                cell = ws.cell(row=summary_row, column=col)
                                if tipo_colors.get(tipo_key):
                                        cell.fill = PatternFill(start_color=tipo_colors[tipo_key], end_color=tipo_colors[tipo_key], fill_type='solid')
                                cell.border = border
                        
                        summary_row += 1

                # Ancho de columnas
                column_widths = [8, 20, 12, 25, 15, 18, 10, 15, 40]
                for col_num, width in enumerate(column_widths, 1):
                        ws.column_dimensions[chr(64 + col_num)].width = width

                # Congelar paneles (filas 1-5)
                ws.freeze_panes = 'A6'

                # SEGUNDA HOJA: LEYENDA DE COLORES
                ws_leyenda = wb.create_sheet('Leyenda')
                
                # Título de la leyenda
                ws_leyenda.merge_cells('A1:D1')
                title_leyenda = ws_leyenda['A1']
                title_leyenda.value = '🎨 GUÍA DE COLORES Y TIPOS DE MOVIMIENTO'
                title_leyenda.font = Font(bold=True, size=14, color='FFFFFF')
                title_leyenda.fill = PatternFill(start_color='1565C0', end_color='1565C0', fill_type='solid')
                title_leyenda.alignment = Alignment(horizontal='center', vertical='center')
                ws_leyenda.row_dimensions[1].height = 25

                # Encabezados de la leyenda
                header_fill_legend = PatternFill(start_color='4A90E2', end_color='4A90E2', fill_type='solid')
                header_font_legend = Font(bold=True, color='FFFFFF', size=11)
                border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
                
                legend_headers = ['Tipo Movimiento', 'Color', 'Descripción', 'Cuándo se usa']
                for col_num, header in enumerate(legend_headers, 1):
                        cell = ws_leyenda.cell(row=3, column=col_num, value=header)
                        cell.fill = header_fill_legend
                        cell.font = header_font_legend
                        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                        cell.border = border
                ws_leyenda.row_dimensions[3].height = 25

                # Datos de la leyenda
                leyenda_data = [
                        ('Agregar', 'E8F5E9', '🟢 Verde claro', 'Cuando se añade stock al inventario (compras, devoluciones, ajustes positivos)'),
                        ('Restar', 'FFEBEE', '🔴 Rojo claro', 'Cuando se reduce stock del inventario (ajustes negativos, merma)'),
                        ('Venta', 'FFF3E0', '🟠 Naranja claro', 'Cuando se realiza una venta y se decuenta del stock'),
                        ('Producto Creado', 'E3F2FD', '🔵 Azul claro', 'Cuando se crea un nuevo producto en el sistema'),
                        ('Producto Eliminado', 'F3E5F5', '🟣 Púrpura claro', 'Cuando se elimina un producto permanentemente'),
                        ('Cambio de estado', 'FFF9C4', '🟡 Amarillo claro', 'Cuando se cambia el estado (Apto→Defectuoso, etc)'),
                        ('Reclamo a proveedor', 'E0F2F1', '🏞️ Teal claro', 'Cuando se envía producto defectuoso al proveedor'),
                        ('Dado de baja', 'ECEFF1', '⚫ Gris azulado', 'Cuando se elimina definitivamente del inventario'),
                        ('Devolución recibida', 'C8E6C9', '💚 Verde intenso', 'Cuando un cliente devuelve un producto'),
                ]
                
                row = 4
                for tipo_display, color_hex, color_emoji, descripcion_uso in leyenda_data:
                        # Columna 1: Tipo de movimiento
                        cell1 = ws_leyenda.cell(row=row, column=1, value=tipo_display)
                        cell1.border = border
                        cell1.alignment = Alignment(horizontal='left', wrap_text=True)
                        cell1.font = Font(bold=True)
                        
                        # Columna 2: Color visual
                        cell2 = ws_leyenda.cell(row=row, column=2, value=color_emoji)
                        cell2.border = border
                        cell2.fill = PatternFill(start_color=color_hex, end_color=color_hex, fill_type='solid')
                        cell2.alignment = Alignment(horizontal='center', vertical='center')
                        
                        # Columna 3: Descripción
                        cell3 = ws_leyenda.cell(row=row, column=3, value=color_emoji)
                        cell3.border = border
                        cell3.alignment = Alignment(horizontal='center', vertical='center')
                        cell3.fill = PatternFill(start_color=color_hex, end_color=color_hex, fill_type='solid')
                        cell3.font = Font(size=14)
                        
                        # Columna 4: Cuándo se usa
                        cell4 = ws_leyenda.cell(row=row, column=4, value=descripcion_uso)
                        cell4.border = border
                        cell4.alignment = Alignment(horizontal='left', wrap_text=True)
                        
                        # Colores de fondo
                        cell1.fill = PatternFill(start_color=color_hex, end_color=color_hex, fill_type='solid')
                        cell4.fill = PatternFill(start_color=color_hex, end_color=color_hex, fill_type='solid')
                        
                        row += 1

                # Ancho de columnas de la leyenda
                ws_leyenda.column_dimensions['A'].width = 22
                ws_leyenda.column_dimensions['B'].width = 16
                ws_leyenda.column_dimensions['C'].width = 12
                ws_leyenda.column_dimensions['D'].width = 50

                output = BytesIO()
                wb.save(output)
                output.seek(0)
                response = HttpResponse(output.getvalue(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
                filename = f'Movimientos_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
                response['Content-Disposition'] = f'attachment; filename="{filename}"'
                try:
                        from urllib.parse import quote
                        response['Content-Disposition'] += f"; filename*=UTF-8''{quote(filename)}"
                except Exception:
                        pass
                return response


class ImagenProductoViewSet(viewsets.ModelViewSet):
        queryset = ImagenProducto.objects.all()
        serializer_class = ImagenProductoSerializer
        permission_classes = [IsAuthenticated, IsEmpleadoReadOnly]
        
        def create(self, request, *args, **kwargs):
                """Crear nueva imagen de producto."""
                if _user_is_empleado(request.user) and not request.user.is_superuser:
                        return Response({'detail': 'No autorizado'}, status=status.HTTP_403_FORBIDDEN)
                
                serializer = self.get_serializer(data=request.data)
                serializer.is_valid(raise_exception=True)
                self.perform_create(serializer)
                return Response(serializer.data, status=status.HTTP_201_CREATED)
        
        @action(detail=False, methods=['get'])
        def por_producto(self, request):
                """Obtener imágenes de un producto específico."""
                producto_id = request.query_params.get('producto_id')
                if not producto_id:
                        return Response({'error': 'producto_id requerido'}, status=status.HTTP_400_BAD_REQUEST)
                
                try:
                        producto = Producto.objects.get(id=producto_id)
                except Producto.DoesNotExist:
                        return Response({'error': 'Producto no encontrado'}, status=status.HTTP_404_NOT_FOUND)
                
                imagenes = ImagenProducto.objects.filter(producto=producto)
                serializer = self.get_serializer(imagenes, many=True)
                return Response(serializer.data)
        
        @action(detail=True, methods=['post'])
        def marcar_principal(self, request, pk=None):
                """Marcar una imagen como principal."""
                if _user_is_empleado(request.user) and not request.user.is_superuser:
                        return Response({'detail': 'No autorizado'}, status=status.HTTP_403_FORBIDDEN)
                
                imagen = self.get_object()
                # Desmarcar las demás imágenes del mismo producto
                ImagenProducto.objects.filter(producto=imagen.producto).update(es_principal=False)
                # Marcar esta como principal
                imagen.es_principal = True
                imagen.save()
                
                return Response(self.get_serializer(imagen).data)