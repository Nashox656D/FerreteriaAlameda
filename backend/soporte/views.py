from rest_framework import viewsets, status, permissions
from rest_framework.decorators import action
from rest_framework.response import Response
from django.http import HttpResponse
from django.contrib.auth.models import User
from django.utils import timezone
from .models import Ticket, Instalacion, SolicitudRecuperacion
from .serializers import TicketSerializer, InstalacionSerializer, SolicitudRecuperacionSerializer
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime
from io import BytesIO
import os
from rest_framework import permissions

class TicketViewSet(viewsets.ModelViewSet):
        queryset = Ticket.objects.all()
        serializer_class = TicketSerializer
        
        def get_queryset(self):
                if self.request.user.is_superuser:
                        return Ticket.objects.all()
                elif self.request.user.is_authenticated:
                        if hasattr(self.request.user, 'empleado'):
                                emp = self.request.user.empleado
                                if emp.cargo in ['Recursos Humanos', 'Admin']:
                                        return Ticket.objects.all()
                                else:
                                        return Ticket.objects.filter(empleado=emp)
                        elif hasattr(self.request.user, 'cliente'):
                                return Ticket.objects.filter(cliente=self.request.user.cliente)
                return Ticket.objects.none()
        
        def perform_create(self, serializer):
                from rest_framework.exceptions import PermissionDenied
                
                if self.request.user.is_superuser:
                        raise PermissionDenied('Los administradores no pueden crear tickets. Solo pueden verlos.')
                
                if self.request.user.is_authenticated:
                        if hasattr(self.request.user, 'empleado'):
                                emp = self.request.user.empleado
                                if emp.cargo in ['Admin', 'Recursos Humanos']:
                                        raise PermissionDenied('Los administradores y RR.HH. no pueden crear tickets. Solo pueden verlos.')
                                serializer.save(creado_por=self.request.user, empleado=emp)
                        elif hasattr(self.request.user, 'cliente'):
                                serializer.save(creado_por=self.request.user, cliente=self.request.user.cliente)
                        else:
                                serializer.save(creado_por=self.request.user)
                else:
                        serializer.save()
        
        @action(detail=False, methods=['get'])
        def descargar_excel(self, request):
                tickets = self.get_queryset()
                
                wb = Workbook()
                ws = wb.active
                ws.title = "Tickets"
                
                header_fill = PatternFill(start_color="4A90E2", end_color="4A90E2", fill_type="solid")
                header_font = Font(bold=True, color="FFFFFF")
                border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
                
                headers = ['ID', 'Asunto', 'Descripción', 'Estado', 'Cliente', 'Empleado', 'Fecha Creación']
                for col_num, header in enumerate(headers, 1):
                        cell = ws.cell(row=1, column=col_num, value=header)
                        cell.fill = header_fill
                        cell.font = header_font
                        cell.alignment = Alignment(horizontal='center', vertical='center')
                        cell.border = border
                
                for row_num, ticket in enumerate(tickets, 2):
                        cliente_nombre = ticket.cliente.nombre if ticket.cliente else '—'
                        empleado_nombre = ticket.empleado.nombre if ticket.empleado else '—'
                        fecha = ticket.fecha_creacion.strftime('%Y-%m-%d %H:%M') if ticket.fecha_creacion else '—'
                        
                        row_data = [
                                ticket.id,
                                ticket.asunto,
                                ticket.descripcion,
                                ticket.estado or 'Abierto',
                                cliente_nombre,
                                empleado_nombre,
                                fecha
                        ]
                        for col_num, value in enumerate(row_data, 1):
                                cell = ws.cell(row=row_num, column=col_num, value=value)
                                cell.border = border
                
                column_widths = [8, 30, 40, 15, 25, 25, 20]
                for col_num, width in enumerate(column_widths, 1):
                        ws.column_dimensions[ws.cell(row=1, column=col_num).column_letter].width = width
                
                output = BytesIO()
                wb.save(output)
                output.seek(0)
                
                response = HttpResponse(output.getvalue(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
                response['Content-Disposition'] = f'attachment; filename="Tickets_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx"'
                return response

class InstalacionViewSet(viewsets.ModelViewSet):
        queryset = Instalacion.objects.all()
        serializer_class = InstalacionSerializer

class SolicitudRecuperacionViewSet(viewsets.ModelViewSet):
        queryset = SolicitudRecuperacion.objects.all()
        serializer_class = SolicitudRecuperacionSerializer
        
        def get_permissions(self):
                if self.action in ['aprobar', 'rechazar', 'cambiar_password', 'list', 'retrieve']:
                        return [permissions.IsAdminUser()]
                return [permissions.AllowAny()]
        
        def get_queryset(self):
                # Solo admin puede ver todas las solicitudes
                if self.request.user.is_superuser:
                        return SolicitudRecuperacion.objects.all().order_by('-fecha_solicitud')
                elif self.request.user.is_authenticated and hasattr(self.request.user, 'empleado'):
                        emp = self.request.user.empleado
                        if emp.cargo in ['Recursos Humanos', 'Admin']:
                                return SolicitudRecuperacion.objects.all().order_by('-fecha_solicitud')
                return SolicitudRecuperacion.objects.none()
        
        def perform_create(self, serializer):
                username = serializer.validated_data.get('username')
                try:
                        user = User.objects.get(username=username)
                        serializer.save(user=user)
                except User.DoesNotExist:
                        serializer.save(user=None)
        
        @action(detail=True, methods=['post'], permission_classes=[permissions.IsAdminUser])
        def aprobar(self, request, pk=None):
                solicitud = self.get_object()
                solicitud.estado = 'APROBADA'
                solicitud.fecha_respuesta = timezone.now()
                solicitud.atendida_por = request.user
                solicitud.save()
                
                return Response({'mensaje': 'Solicitud aprobada exitosamente'})
        
        @action(detail=True, methods=['post'], permission_classes=[permissions.IsAdminUser])
        def cambiar_password(self, request, pk=None):
                solicitud = self.get_object()
                nueva_password = request.data.get('nueva_contraseña') or request.data.get('nueva_password')

                if not nueva_password:
                        return Response({'error': 'Debe proporcionar una nueva contraseña'}, status=status.HTTP_400_BAD_REQUEST)

                # If the solicitud already has a user FK, use it. Otherwise try to resolve
                # the user now by the stored username (in case it was not found at creation).
                user = solicitud.user
                if user is None and solicitud.username:
                        try:
                                user = User.objects.get(username=solicitud.username)
                        except User.DoesNotExist:
                                user = None

                if user:
                        try:
                                # Optionally validate password strength using Django validators
                                from django.contrib.auth.password_validation import validate_password
                                validate_password(nueva_password, user=user)
                        except Exception as e:
                                return Response({'error': str(e)}, status=status.HTTP_400_BAD_REQUEST)

                        user.set_password(nueva_password)
                        user.save()
                        solicitud.delete()

                        return Response({'mensaje': 'Contraseña actualizada exitosamente y solicitud eliminada'})
                else:
                        return Response({'error': 'Usuario no encontrado'}, status=status.HTTP_404_NOT_FOUND)
        
        @action(detail=True, methods=['post'], permission_classes=[permissions.IsAdminUser])
        def rechazar(self, request, pk=None):
                solicitud = self.get_object()
                solicitud.estado = 'RECHAZADA'
                solicitud.fecha_respuesta = timezone.now()
                solicitud.atendida_por = request.user
                solicitud.save()
                
                return Response({'mensaje': 'Solicitud rechazada'})


 
